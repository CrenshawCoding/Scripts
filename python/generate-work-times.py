import datetime
import sys
import string
import os
import random
from openpyxl import *
from shutil import copyfile
#for export, not working currently:
#from win32com import client 
#from pywintypes import com_error



#--- Set these values to your desire ---#
# contains work times
# workTimes[0][0] = monday start
# workTimes[0][1] = monday end
# 11 = 11:00, 15 = 15:00 etc. Use integers. 0 if you are not working that day.
workTimes = [[0, 0], [10, 15], [0, 0], [11,15], [0, 0], [0,0], [0,0]]
myName = "Harlos" # your name, duh
TEMPLATE_NAME = "Stundennnachweis_Harlos_Template.xlsx" # Path to the template
SHEET_NAME = "Tabelle1" # Name of the sheet to write into (in German usually "Tabelle1", look into your template)
#---------------------------------------#



if(len(sys.argv) > 4 | len(sys.argv) < 2):
    raise ValueError('Usage: python generate-work-times.py <dd-mm-yyyy> [-r <offset in minutes>]')
    
startDayParam = sys.argv[1]
addRandomness = False
maxOffset = 0
if(len(sys.argv) == 4):
    if(sys.argv[2] == "-r"):
        maxOffset = sys.argv[3]
        addRandomness = True

# get the start day from command line param:
startDay = datetime.date(int(startDayParam[6:10]), int(startDayParam[3:5]), int(startDayParam[0:2]))
curMonth = startDay.month
curYear = startDay.year
oneDayDelta = datetime.timedelta(days=1)


COPY_NAME = "Stundenzettel_{0}_{1}-{2}.xlsx".format(myName, curMonth, curYear)


#copy the template into a new copy
copyfile(TEMPLATE_NAME, COPY_NAME) 

wb = load_workbook(COPY_NAME)
sheet = wb[SHEET_NAME]



row = 13
column = 2
while startDay.month == curMonth:
    sheet.cell(row, column).value = startDay.strftime("%d.%m.%Y") # write date
    startTime = 0
    endTime = 0
    if addRandomness:
        offsetMinutes = (random.randint(-(int(maxOffset)), int(maxOffset)))
        offsetMinutes = offsetMinutes - (offsetMinutes % 10)
        startTime = datetime.timedelta(hours=workTimes[startDay.weekday()][0], minutes=offsetMinutes)
        endTime = datetime.timedelta(hours=workTimes[startDay.weekday()][1], minutes=offsetMinutes)
        
    else:
        startTime = datetime.timedelta(hours=workTimes[startDay.weekday()][0])
        endTime = datetime.timedelta(hours=workTimes[startDay.weekday()][1])
        
    sheet.cell(row, column + 1).value = startTime if workTimes[startDay.weekday()][0] != 0 else ""
    sheet.cell(row, column + 2).value = endTime if workTimes[startDay.weekday()][0] != 0 else ""
    
    startDay += oneDayDelta
    row += 1

print("saving", os.getcwd() + os.path.sep + COPY_NAME, "...")
wb.save(os.getcwd() + os.path.sep + COPY_NAME)

# convert to pdf, not working
"""
print("converting: ", os.getcwd() + os.path.sep + COPY_NAME, "...")
app = client.DispatchEx("Excel.Application")
app.Interactive = False
app.Visible = False
Workbook = app.Workbooks.Open(os.getcwd() + os.path.sep + COPY_NAME)
try:
    file_name = os.path.splitext(os.getcwd() + os.path.sep + COPY_NAME)[0]
    output_file = os.getcwd() + os.path.sep + file_name + ".pdf"
    wb.WorkSheets(1).Select()
    Workbook.ActiveSheet.ExportAsFixedFormat(0, "C:\\Users\\leand\\OneDrive\\Arbeit\\test.pdf")
except com_error as e:
    print("Failed to convert in PDF format.Please confirm environment meets all the requirements  and try again")
    print(str(e))
finally:
    Workbook.Close()
    app.Exit()
"""
print("Done")