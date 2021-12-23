import argparse
import wordcounter
import os
import openpyxl
import datetime
from datetime import date
from datetime import datetime, timedelta

DATE_COLUMN = 1
WORDS_COLUMN = 2
CHARACTERS_COLUMN = 3
TIME_FORMAT = "%d.%m.%Y"

tex_file_path = ".\\Loss Detection and Recovery mechanisms in the QUIC protocol.tex"
tex_file = open(tex_file_path)
content = tex_file.read()
words = content.split()
content_no_newline = content.replace("\n", "")
characters = len(content_no_newline)
print(len(words), len(content_no_newline))

xlsx_file_path = ".\\{0}-progress.xlsx".format(os.path.basename(tex_file.name))
column_names = ["Date", "Words", "Characters"]

#Set up the File if it does not exist
if(not os.path.isfile(xlsx_file_path)):
    wb = openpyxl.Workbook()
    i = 1
    for name in column_names:
        wb.active.cell(1, i).value = name
        i += 1
    wb.save(xlsx_file_path)
    
wb = openpyxl.load_workbook(xlsx_file_path)
sheet = wb.active
row_iter = 2
today = date.today().strftime(TIME_FORMAT)
cur_date_cell = sheet.cell(row_iter, DATE_COLUMN)
cell_datetime = None
while(cur_date_cell.value != None):
    cell_datetime = sheet.cell(row_iter, DATE_COLUMN).value
    if(cell_datetime.date() == date.today()):
        print("DEBUG")
        break
    row_iter += 1
    cur_date_cell = sheet.cell(row_iter, DATE_COLUMN)

#If the last found date is before yesterday, fill out rows with 0 progress
if cell_datetime and cell_datetime.date() < date.today() - timedelta(days = 1):
    print("Creating entries for missed days from {0} to {1}".format(datetime.strftime(cell_datetime + timedelta(days = 1), TIME_FORMAT), 
        datetime.strftime(date.today() - timedelta(days = 1), TIME_FORMAT)))
    timestamp = cell_datetime + timedelta(days = 1)
    last_found_words = sheet.cell(row_iter - 1, WORDS_COLUMN).value
    last_found_characters = sheet.cell(row_iter - 1, CHARACTERS_COLUMN).value
    while timestamp.date() < date.today():
        sheet.cell(row_iter, DATE_COLUMN).value = datetime.strftime(timestamp, TIME_FORMAT)
        sheet.cell(row_iter, WORDS_COLUMN).value = last_found_words
        sheet.cell(row_iter, CHARACTERS_COLUMN).value = last_found_characters
        timestamp += timedelta(days = 1)
        row_iter += 1
else:
    if not cell_datetime:
        sheet.cell(row_iter, DATE_COLUMN).value = datetime.strftime(date.today(), TIME_FORMAT)
        print("Creating new entry for today with {0} words and {1} characters".format(len(words), characters)) 
    else:
        print("DEBUG", row_iter)
        print("Updating today's progress from {0} words and {1} characters to {2} words and {3} characters"
            .format(sheet.cell(row_iter, WORDS_COLUMN).value, sheet.cell(row_iter, CHARACTERS_COLUMN).value, len(words), characters))
    sheet.cell(row_iter, WORDS_COLUMN).value = len(words)
    sheet.cell(row_iter, CHARACTERS_COLUMN).value = characters

wb.save(xlsx_file_path)
